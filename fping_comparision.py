## Importing Library ##

import os
import re
import glob
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Color, Font, colors, PatternFill, Font, Border


## Hardcore Variables ##

var_s1 = "is alive"
var_s2 = "is unreachable"


var_s4 = "for ICMP Echo sent to "
lvar_s4 = len(var_s4)
pre_alive = []
pre_unreachable = []
post_alive = []
post_unreachable = []
icmp_unreachable_from_pre = []
icmp_unreachable_to_pre = []
icmp_unreachable_from_post = []
icmp_unreachable_to_post = []
icmp_unreachable_to_pre_ip_address = []
icmp_unreachable_to_pre_device_name = []
icmp_unreachable_to_post_ip_address = []
icmp_unreachable_to_post_device_name = []
icmp_add_pre_unreachable = []
icmp_add_post_unreachable = []

## Methods ##

def from_text_file_to_variable (fi_text_file) :
	f = open(fi_text_file, 'r')
	fo_variable = f.read()
	f.close()
	return fo_variable
	
def from_list_to_excel_worksheet (fi_list, fi_excel_workbook, fi_excel_worksheet, fi_row, fi_column) :
	wb = load_workbook(fi_excel_workbook)								
	ws = wb.get_sheet_by_name(fi_excel_worksheet)					
	fi_1 = fi_row
	for fi_x1 in fi_list :
		ws.cell(row = fi_1, column = fi_column).value = fi_x1
		fi_1 = fi_1 + 1
	wb.save("Fping Report.xlsx")

def extract_first_ip_from_string (fi_string):
	fi_var_1 = re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', fi_string)
	if (fi_var_1 is not None):
		fo_ip = fi_var_1.group()
		return fo_ip

def autofit_all_columns_of_excel_worksheet_in_workbook (fi_workbook, fi_worksheet):
	import win32com.client as win32
	excel = win32.gencache.EnsureDispatch('Excel.Application')
	wb = excel.Workbooks.Open(fi_workbook)
	ws = wb.Worksheets(fi_worksheet)
	ws.Columns.AutoFit()
	wb.Save()
		
##Input from the User

pre_remove = "y"
post_remove = "y"

#### Code ####

## Open File ##
var_fpre = from_text_file_to_variable ("fpre.txt")
var_fpost = from_text_file_to_variable ("fpost.txt")

## Fetching Required String/Substring
#var_expression = "\n(.*?)\n"
#var_re_1 = re.compile(var_expression, re.DOTALL |  re.IGNORECASE).findall(var_fpre)
#var_re_2 = re.compile(var_expression, re.DOTALL |  re.IGNORECASE).findall(var_fpost)

var_re_1 = var_fpre.split("\n")
var_re_2 = var_fpost.split("\n")

## Creating Pre List
for var_x_1 in var_re_1 :
	var_result_1 = var_x_1.find(var_s1) 
	var_result_2 = var_x_1.find(var_s2)
	if (pre_remove == "n"):
		if (var_result_1 != -1):
			pre_alive.append(var_x_1[:var_result_1 -1])
		if (var_result_2 != -1):
			pre_unreachable.append(var_x_1[:var_result_2 -1])
	elif (pre_remove == "y"):
		if (var_result_1 != -1):
			pre_alive.append(var_x_1[:var_result_1 -1].replace(".ms.com", ""))
		if (var_result_2 != -1):
			pre_unreachable.append(var_x_1[:var_result_2 -1].replace(".ms.com", ""))
			
## Creating Post List		
for var_x_2 in var_re_2 :
	var_result_1 = var_x_2.find(var_s1)
	#print (var_result_1)
	var_result_2 = var_x_2.find(var_s2)
	if (post_remove == "n"):
		if (var_result_1 != -1):
			post_alive.append(var_x_2[:var_result_1 -1])
		if (var_result_2 != -1):
			post_unreachable.append(var_x_2[:var_result_2 -1])
	elif (post_remove == "y"):
		if (var_result_1 != -1):
			post_alive.append(var_x_2[:var_result_1 -1].replace(".ms.com", ""))
		if (var_result_2 != -1):
			post_unreachable.append(var_x_2[:var_result_2 -1].replace(".ms.com", ""))
	
	
## ICMP Unreachable
#Pre List
for var_x_3 in var_re_1 :
	#var_result_1 = var_x_3.find(var_s3)
	var_result_2 = var_x_3.find(var_s4)
	if (var_result_2 != -1):
		from_ip = extract_first_ip_from_string (var_x_3)		
		icmp_unreachable_from_pre.append(from_ip)
		if (pre_remove == "n"):
			icmp_unreachable_to_pre.append(var_x_3[var_result_2 + lvar_s4:])
		elif (pre_remove == "y"):
			icmp_unreachable_to_pre.append(var_x_3[var_result_2 + lvar_s4:].replace(".ms.com", ""))
		
		
for var_x_5 in icmp_unreachable_to_pre :
	var_result_1 = var_x_5.split( )
	lvar_result_1 = len(var_result_1)
	if (pre_remove == "n"):
		icmp_add_pre_unreachable.append(var_result_1[0])
		if lvar_result_1 == 1 :
			icmp_unreachable_to_pre_device_name.append(" ")
			icmp_unreachable_to_pre_ip_address.append(var_result_1[0])
		elif lvar_result_1 == 2 :
			icmp_unreachable_to_pre_device_name.append(var_result_1[0])
			icmp_unreachable_to_pre_ip_address.append(var_result_1[1][1:-1])
		else :
			print ("Something Suspicious in ICMP Pre Logs")
	elif (pre_remove == "y"):	
		icmp_add_pre_unreachable.append(var_result_1[0].replace(".ms.com", ""))
		if lvar_result_1 == 1 :
			icmp_unreachable_to_pre_device_name.append(" ")
			icmp_unreachable_to_pre_ip_address.append(var_result_1[0].replace(".ms.com", ""))
		elif lvar_result_1 == 2 :
			icmp_unreachable_to_pre_device_name.append(var_result_1[0].replace(".ms.com", ""))
			icmp_unreachable_to_pre_ip_address.append(var_result_1[1][1:-1].replace(".ms.com", ""))
		else :
			print ("Something Suspicious in ICMP Pre Logs")
		
	

#Post List
for var_x_4 in var_re_2 :
	#var_result_1 = var_x_4.find(var_s3)
	var_result_2 = var_x_4.find(var_s4)
	if (var_result_2 != -1):
		from_ip = extract_first_ip_from_string (var_x_4)
		icmp_unreachable_from_post.append(from_ip)
		if (post_remove == "n"):
			icmp_unreachable_to_post.append(var_x_4[var_result_2 + lvar_s4:])
		elif (post_remove == "y"):
			icmp_unreachable_to_post.append(var_x_4[var_result_2 + lvar_s4:].replace(".ms.com", ""))
		
	
for var_x_6 in icmp_unreachable_to_post :
	var_result_1 = var_x_6.split( )
	lvar_result_1 = len(var_result_1)
	if (post_remove == "n"):
		icmp_add_post_unreachable.append(var_result_1[0])
		if lvar_result_1 == 1 :
			icmp_unreachable_to_post_device_name.append(" ")
			icmp_unreachable_to_post_ip_address.append(var_result_1[0])
		elif lvar_result_1 == 2 :
			icmp_unreachable_to_post_device_name.append(var_result_1[0])
			icmp_unreachable_to_post_ip_address.append(var_result_1[1][1:-1])
		else :
			print ("Something Suspicious in ICMP Post Logs")
	elif (post_remove == "y"):
		icmp_add_post_unreachable.append(var_result_1[0].replace(".ms.com", ""))
		if lvar_result_1 == 1 :
			icmp_unreachable_to_post_device_name.append(" ")
			icmp_unreachable_to_post_ip_address.append(var_result_1[0].replace(".ms.com", ""))
		elif lvar_result_1 == 2 :
			icmp_unreachable_to_post_device_name.append(var_result_1[0].replace(".ms.com", ""))
			icmp_unreachable_to_post_ip_address.append(var_result_1[1][1:-1].replace(".ms.com", ""))
		else :
			print ("Something Suspicious in ICMP Post Logs")
	

	
	
## Diff Check ##
'''
print ("Alive : The one alive in pre check, however not alive in post check")
print (list(set(pre_alive) - set(post_alive)))

print ("Alive : The one alive in post check, however not alive in pre check")
print (list(set(post_alive) - set(pre_alive)))

print ("Unreachable : The one unreachable in pre check, however not unreachable in post check")
print (list(set(pre_unreachable) - set(post_unreachable)))

print ("Unreachable : The one unreachable in post check, however not unreachable in pre check")
print (list(set(post_unreachable) - set(pre_unreachable)))
'''


## Adding to Pre and Post Unreachable (ICMP)

pre_unreachable = pre_unreachable + icmp_add_pre_unreachable
post_unreachable = post_unreachable + icmp_add_post_unreachable


## Union of Pre and Post
pre_list = [pre_alive, pre_unreachable]
pre_union = set().union(*pre_list)

post_list = [post_alive, post_unreachable]
post_union = set().union(*post_list)


## Missing in precheck

list_premiss = list(set(post_union) - set(pre_union))

## Missing in postcheck

list_postmiss = list(set(pre_union) - set(post_union))

#The one alive in precheck and unreachable in post check
list_au = [val for val in pre_alive if val in post_unreachable]

#The one alive in precheck and missing in post check
#list_am = list(set(pre_alive) - set(post_union))
list_am = [val for val in pre_alive if val in list_postmiss]

#The one unreachable in precheck and alive in post check
list_ua = [val for val in pre_unreachable if val in post_alive]

#The one unreachable in precheck and missing in post check
#list_um = list(set(pre_unreachable) - set(post_union))
list_um = [val for val in pre_unreachable if val in list_postmiss]

#The one missing in precheck and alive in post check
#list_ma = list(set(post_alive) - set(pre_union))
list_ma = [val for val in list_premiss if val in post_alive]

#The one missing in precheck and unreachable in post check
#list_mu = list(set(post_unreachable) - set(pre_union))
list_mu = [val for val in list_premiss if val in post_unreachable]


from_list_to_excel_worksheet(list_au, "Report Format.xlsx", "Sheet1", 9, 1)
from_list_to_excel_worksheet(list_am, "Fping Report.xlsx", "Sheet1", 9, 2)
from_list_to_excel_worksheet(list_ua, "Fping Report.xlsx", "Sheet1", 9, 3)
from_list_to_excel_worksheet(list_um, "Fping Report.xlsx", "Sheet1", 9, 4)
from_list_to_excel_worksheet(list_ma, "Fping Report.xlsx", "Sheet1", 9, 5)
from_list_to_excel_worksheet(list_mu, "Fping Report.xlsx", "Sheet1", 9, 6)


#autofit_all_columns_of_excel_worksheet_in_workbook (os.getcwd() + "\\Fping Report.xlsx", "Sheet1")

